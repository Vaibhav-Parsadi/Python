'''
import openpyxl
wb=openpyxl.load_workbook("C:\\Users\\Vaibhav\\Desktop\\pl.xlsx")
'''
'''
print(type(wb))
sheets=wb.sheetnames
print(wb.active.title)
sh1=wb['Sheet1']
print(type(sh1))
data=sh1['A2'].value #way to access dat
print("A2 in data variable is store with value",data)
d2=print("d2 is",wb["Sheet1"]["B2"].value) #This way also we can access data
#we are treating wb as list with names and index specified

#option 3 using rows,column index
print("option 3 ",sh1.cell(4,2).value)
'''
'''
#--------------------------------------------------------------------------------------------------------------------
#same thing but getting all the data using loop with rows and column function
#--------------------------------------------------------------------------------------------------------------------
sh1=wb['Sheet1']    #sheet1 of wb is stored in sh1
print(type(sh1))
row=sh1.max_row                 #this fn gives total no. of rows in object sh1
column=sh1.max_column           #this fn gives total no. of column in object sh1
print("rows-",row,"Columns are-",column)
roll=[];name=[];marks=[]
for i in range(1,row+1):                                       #here loop goes till no of rows in sh1 +1
    for j in range(1,column+1):                                #here loop goes till no of columns in sh1 +1
        if j==1:
            name.append(sh1.cell(i,j).value)                   #appending in list
        if j==2:
            roll.append(sh1.cell(i, j).value)
        if j==3:
            marks.append(sh1.cell(i, j).value)
    print()
c=list(zip(name,marks))
d=dict(zip(roll,c))
print("zip lists are ",c[0])
print(d)
sh1.cell(row=9,column=1,value="Zubir")                          #we can add new record in sheet by using this function
sh1.cell(row=9,column=2,value=8)                                #diff values are added to their resp. columns
sh1.cell(row=9,column=3,value=99)
wb.save("C:\\Users\\Vaibhav\\Desktop\\pl.xlsx")                 #to save the file we need filename/path within wb.save()
'''
#----------------------------------------------------------------------------------------------------------------------
# NOW WE WILL CREATE TABLE FROM EMPTY SHEET
#----------------------------------------------------------------------------------------------------------------------
from openpyxl import Workbook
wb=Workbook()                    #object of Workbook class created
print(wb.active.title)
print(wb.sheetnames)
wb['Sheet'].title="Creation"     #We can set the title by this function
s1=wb.active
s1['A1'].value="Name"            #Value for A1 cell is set as 'Name'
s1['B1'].value="Roll"
s1['C1'].value="City"
rows=int(input("enter no of records you want "))
column=s1.max_column
print("columns are ",column)
for i in range(1,rows+1):
    for j in range(1,column+1):
        s1.cell(row=i,column=j,value=input("enter input"))

for i in range(1,rows+1):
    for j in range(1,column+1):
        print(s1.cell(row=i,column=j).value,end=" ")
    print()

wb.save("C:\\Users\\Vaibhav\\Desktop\\creation.xlsx")
