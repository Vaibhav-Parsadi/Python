import openpyxl
w=openpyxl.load_workbook("C:\\Users\\Vaibhav\\Desktop\\creation.xlsx")
print(type(w))
print(w.active)
s1=w.active
print(type(s1))

row=s1.max_row
column=s1.max_column
print(row,"   ",column)
s1.cell(row=5,column=1,value=5)
for i in range(1,row+1):
    for j in range(1,column+1):
        print(s1.cell(i,j).value,end=" ")
    print()
